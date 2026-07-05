from django.contrib.auth import get_user_model
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

from apps.teams.models import Team, TeamMember
from apps.teams.services import create_daily_logs_for_member, create_innovation_brief

User = get_user_model()

SHEET_TEACHERS = "老师 Teachers"
SHEET_STUDENTS = "学生 Students"
SHEET_TEAMS = "队伍 Teams"
SHEET_MEMBERS = "队伍成员 Team Members"

HEADER_ALIASES = {
    "username*": "username",
    "password*": "password",
    "display_name*": "display_name",
    "email*": "email",
    "team_name*": "team_name",
    "project_name*": "project_name",
    "challenge_category*": "challenge_category",
    "teacher_username*": "teacher_username",
    "student_username*": "student_username",
}


def norm_header(cell):
    if cell is None:
        return ""
    key = str(cell).strip()
    return HEADER_ALIASES.get(key, key)


def cell_str(value):
    if value is None:
        return ""
    return str(value).strip()


def read_sheet(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        raise CommandError(f"Missing sheet: {sheet_name}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [norm_header(h) for h in rows[0]]
    data = []
    for row_num, row in enumerate(rows[1:], start=2):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        item = {headers[i]: cell_str(row[i]) if i < len(row) else "" for i in range(len(headers))}
        item["_row"] = row_num
        data.append(item)
    return data


class Command(BaseCommand):
    help = "Import teachers, students, teams, and members from Excel template"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Path to filled Excel file")
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Validate only, do not write to database",
        )
        parser.add_argument(
            "--skip-demo-cleanup",
            action="store_true",
            help="Do not delete seed_demo accounts (teacher_demo, student01-05)",
        )

    def handle(self, *args, **options):
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise CommandError("Install openpyxl: pip install openpyxl") from exc

        path = options["xlsx_path"]
        dry_run = options["dry_run"]

        wb = load_workbook(path, data_only=True)
        teachers = read_sheet(wb, SHEET_TEACHERS)
        students = read_sheet(wb, SHEET_STUDENTS)
        teams = read_sheet(wb, SHEET_TEAMS)
        members = read_sheet(wb, SHEET_MEMBERS)
        wb.close()

        self.stdout.write(f"Teachers: {len(teachers)}, Students: {len(students)}, Teams: {len(teams)}, Members: {len(members)}")

        self._validate(teachers, students, teams, members)

        if dry_run:
            self.stdout.write(self.style.SUCCESS("Dry run OK — no changes made."))
            return

        with transaction.atomic():
            if not options["skip_demo_cleanup"]:
                self._cleanup_demo()
            teacher_map = self._import_teachers(teachers)
            student_map = self._import_students(students)
            team_map = self._import_teams(teams, teacher_map)
            self._import_members(members, team_map, student_map)

        self.stdout.write(self.style.SUCCESS("Import complete."))

    def _validate(self, teachers, students, teams, members):
        required_teacher = {"username", "password", "display_name", "email"}
        required_student = {"username", "password", "display_name", "email"}
        required_team = {"team_name", "project_name", "challenge_category", "teacher_username"}
        required_member = {"team_name", "student_username"}

        usernames = set()
        emails = set()

        for row in teachers + students:
            for field in ("username", "email"):
                val = row.get(field, "")
                if not val:
                    raise CommandError(f"Row {row['_row']}: missing {field}")
            if row["username"] in usernames:
                raise CommandError(f"Duplicate username: {row['username']}")
            if row["email"] in emails:
                raise CommandError(f"Duplicate email: {row['email']}")
            usernames.add(row["username"])
            emails.add(row["email"])

        for row in teachers:
            if not all(row.get(f) for f in required_teacher):
                raise CommandError(f"Teacher row {row['_row']}: missing required fields")

        for row in students:
            if not all(row.get(f) for f in required_student):
                raise CommandError(f"Student row {row['_row']}: missing required fields")

        teacher_usernames = {t["username"] for t in teachers}

        team_names = set()
        for row in teams:
            if not all(row.get(f) for f in required_team):
                raise CommandError(f"Team row {row['_row']}: missing required fields")
            if row["team_name"] in team_names:
                raise CommandError(f"Duplicate team_name: {row['team_name']}")
            if row["teacher_username"] not in teacher_usernames:
                raise CommandError(
                    f"Team '{row['team_name']}': teacher_username '{row['teacher_username']}' not in teachers sheet"
                )
            team_names.add(row["team_name"])

        student_usernames = {s["username"] for s in students}
        member_counts = {}
        assigned_students = set()

        for row in members:
            if not all(row.get(f) for f in required_member):
                raise CommandError(f"Member row {row['_row']}: missing required fields")
            if row["team_name"] not in team_names:
                raise CommandError(f"Member row {row['_row']}: unknown team '{row['team_name']}'")
            if row["student_username"] not in student_usernames:
                raise CommandError(
                    f"Member row {row['_row']}: unknown student '{row['student_username']}'"
                )
            if row["student_username"] in assigned_students:
                raise CommandError(f"Student '{row['student_username']}' assigned to multiple teams")
            assigned_students.add(row["student_username"])
            member_counts[row["team_name"]] = member_counts.get(row["team_name"], 0) + 1

        for team_name, count in member_counts.items():
            if count > TeamMember.MAX_MEMBERS:
                raise CommandError(f"Team '{team_name}' has {count} members (max {TeamMember.MAX_MEMBERS})")

    def _cleanup_demo(self):
        demo_users = User.objects.filter(username__in=["teacher_demo"] + [f"student0{i}" for i in range(1, 6)])
        demo_teams = Team.objects.filter(name="Team Mimo")
        count_u, _ = demo_users.delete()
        count_t, _ = demo_teams.delete()
        if count_u or count_t:
            self.stdout.write(f"Removed demo data: {count_u} users, {count_t} teams")

    def _import_teachers(self, rows):
        mapping = {}
        for row in rows:
            user, created = User.objects.update_or_create(
                username=row["username"],
                defaults={
                    "email": row["email"],
                    "role": User.Role.TEACHER,
                    "display_name": row["display_name"],
                    "school": row.get("school", ""),
                    "is_active": True,
                },
            )
            user.set_password(row["password"])
            user.save()
            mapping[row["username"]] = user
            self.stdout.write(f"{'Created' if created else 'Updated'} teacher: {user.username}")
        return mapping

    def _import_students(self, rows):
        mapping = {}
        for row in rows:
            user, created = User.objects.update_or_create(
                username=row["username"],
                defaults={
                    "email": row["email"],
                    "role": User.Role.STUDENT,
                    "display_name": row["display_name"],
                    "school": row.get("school", ""),
                    "grade": row.get("grade", ""),
                    "is_active": True,
                },
            )
            user.set_password(row["password"])
            user.save()
            mapping[row["username"]] = user
            self.stdout.write(f"{'Created' if created else 'Updated'} student: {user.username}")
        return mapping

    def _import_teams(self, rows, teacher_map):
        mapping = {}
        for row in rows:
            teacher = teacher_map[row["teacher_username"]]
            team, created = Team.objects.update_or_create(
                name=row["team_name"],
                defaults={
                    "project_name": row["project_name"],
                    "challenge_category": row["challenge_category"],
                    "teacher": teacher,
                    "description": row.get("description", ""),
                },
            )
            create_innovation_brief(team)
            mapping[row["team_name"]] = team
            self.stdout.write(f"{'Created' if created else 'Updated'} team: {team.name}")
        return mapping

    def _import_members(self, rows, team_map, student_map):
        for row in rows:
            team = team_map[row["team_name"]]
            student = student_map[row["student_username"]]
            member, created = TeamMember.objects.update_or_create(
                team=team,
                student=student,
                defaults={"student_role": row.get("student_role", "")},
            )
            create_daily_logs_for_member(team, student)
            self.stdout.write(
                f"{'Added' if created else 'Updated'} member: {student.username} -> {team.name}"
            )
