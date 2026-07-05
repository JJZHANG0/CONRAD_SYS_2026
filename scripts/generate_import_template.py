#!/usr/bin/env python3
"""Generate bulk import Excel template for teachers, students, teams."""
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("Run: pip install openpyxl")

OUT = Path(__file__).resolve().parent.parent / "docs" / "bulk_import_template.xlsx"

HEADERS = {
    "instructions": ["说明项", "内容"],
    "teachers": [
        "username*",
        "password*",
        "display_name*",
        "email*",
        "school",
    ],
    "students": [
        "username*",
        "password*",
        "display_name*",
        "email*",
        "school",
        "grade",
    ],
    "teams": [
        "team_name*",
        "project_name*",
        "challenge_category*",
        "teacher_username*",
        "description",
    ],
    "team_members": [
        "team_name*",
        "student_username*",
        "student_role",
    ],
}

EXAMPLES = {
    "teachers": [
        ["teacher_zhang", "ChangeMe123!", "张老师", "zhang@school.edu.cn", "杭州某某中学"],
        ["teacher_li", "ChangeMe123!", "李老师", "li@school.edu.cn", "杭州某某中学"],
    ],
    "students": [
        ["student_wang01", "ChangeMe123!", "王小明", "wang01@student.edu.cn", "杭州某某中学", "G10"],
        ["student_li02", "ChangeMe123!", "李小红", "li02@student.edu.cn", "杭州某某中学", "G10"],
        ["student_chen03", "ChangeMe123!", "陈小华", "chen03@student.edu.cn", "杭州某某中学", "G11"],
    ],
    "teams": [
        ["Team Alpha", "Smart Pillow", "Health & Nutrition", "teacher_zhang", "智能枕头项目"],
        ["Team Beta", "Eco Bottle", "Energy & Environment", "teacher_li", "环保水杯项目"],
    ],
    "team_members": [
        ["Team Alpha", "student_wang01", "CEO"],
        ["Team Alpha", "student_li02", "CTO"],
        ["Team Alpha", "student_chen03", "CMO"],
        ["Team Beta", "student_wang01", ""],
    ],
}

INSTRUCTIONS = [
    ("模板用途", "批量创建老师账号、学生账号、队伍及队伍成员关系"),
    ("带 * 的列", "必填字段，不能为空"),
    ("username", "登录用户名，全系统唯一，建议英文+数字，如 student_wang01"),
    ("password", "初始登录密码，建议首次登录后修改；导入脚本会加密存储"),
    ("email", "邮箱，全系统唯一，每个账号一个"),
    ("display_name", "显示名称，建议填中文姓名"),
    ("school / grade", "学校、年级，学生建议填写"),
    ("teacher_username", "填写「老师」工作表中的 username，用于绑定带队老师"),
    ("team_name", "队伍名称，全系统唯一；「队伍成员」表通过此字段关联"),
    ("project_name", "项目名称，如 Mimo Smart Pillow"),
    ("challenge_category", "挑战类别，如 Health & Nutrition / Energy & Environment"),
    ("student_role", "队内角色，可选，如 CEO / CTO / CMO / Designer"),
    ("队伍人数", "每支队伍最多 5 名学生"),
    ("学生归属", "每名学生只能加入 1 支队伍"),
    ("填写顺序", "先填老师 → 再填学生 → 再填队伍 → 最后填队伍成员"),
    ("示例数据", "各工作表中的灰色示例行可删除，仅作参考"),
    ("提交方式", "填好后将 Excel 文件发给技术同事，或使用 import 脚本导入"),
]


def style_header(ws, row=1):
    fill = PatternFill("solid", fgColor="2563EB")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_example(ws, start_row, end_row):
    fill = PatternFill("solid", fgColor="F3F4F6")
    for r in range(start_row, end_row + 1):
        for cell in ws[r]:
            cell.fill = fill


def autosize(ws):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[letter].width = min(width, 40)


def main():
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Sheet 1: Instructions
    ws0 = wb.active
    ws0.title = "填写说明"
    ws0.append(HEADERS["instructions"])
    style_header(ws0)
    for row in INSTRUCTIONS:
        ws0.append(list(row))
    autosize(ws0)

    sheets = [
        ("老师 Teachers", "teachers"),
        ("学生 Students", "students"),
        ("队伍 Teams", "teams"),
        ("队伍成员 Team Members", "team_members"),
    ]

    for title, key in sheets:
        ws = wb.create_sheet(title)
        ws.append(HEADERS[key])
        style_header(ws)
        start = 2
        for ex in EXAMPLES[key]:
            ws.append(ex)
        style_example(ws, start, start + len(EXAMPLES[key]) - 1)
        autosize(ws)

    wb.save(OUT)
    print(f"Created: {OUT}")


if __name__ == "__main__":
    main()
